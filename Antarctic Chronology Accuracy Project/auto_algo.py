#import packages
import numpy as np
import scipy
import matplotlib.pyplot as plt
import pandas as pd
import xarray as xr
import os
import yaml
import itertools
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from collections import defaultdict
import sys
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from numpy._core.numeric import indices
from tqdm import tqdm
from functools import lru_cache
from collections import deque
import networkx as nx
from collections import Counter


#set working dir
os.chdir("/Users/quinnmackay/Documents/GitHub/BICC/Antarctic Chronology Accuracy Project")

# load cores
project = 'all_tiepoints'
output_dir = 'table_out/'

# get all link combos
with open(f'/Users/quinnmackay/Documents/GitHub/BICC/Antarctic Chronology Accuracy Project/{project}/parameters.yml') as f:
    data = yaml.safe_load(f)
list_sites = data["list_sites"]
pairs = [f"{a}-{b}" for a, b in itertools.combinations(list_sites, 2)]

error_margin = 0.15
big_error_margin = 0.25

big_table = pd.DataFrame()
all_links_count = {}
all_links_foragesort = {}
all_links_total = {}

for core in list_sites: # loop through each core
    for comparison_core in list_sites: # loop through each core other than the initial load
        pair = f"{core}-{comparison_core}"
        if core != comparison_core and pair in pairs: # make sure not the same core and we skip non-existent linkages
            pair_dir = Path(f'/Users/quinnmackay/Documents/GitHub/BICC/Antarctic Chronology Accuracy Project/{project}/{pair}')

            # Check: directory exists AND contains at least one .txt file
            txt_files = list(pair_dir.glob("*.txt"))
            if not pair_dir.is_dir() or not txt_files:
                continue

            dfs=[] #load all text files into one
            for txt in txt_files:
                df = pd.read_csv(txt, sep="\t", comment="#")
                dfs.append(df)
    
            num_files = len(dfs)
            load_data = pd.concat(dfs, ignore_index=True)
            original_rows = len(load_data)

            drop_rows = []
            drop_rows_merge = set()
            new_merged_rows = []
            for idx, row in load_data.iterrows():

                mask1 = abs(row['depth1'] - load_data['depth1']) <= error_margin
                mask1[idx] = False
                mask2 = abs(row['depth2'] - load_data['depth2']) <= error_margin 
                mask2[idx] = False

                close_points = load_data[mask1 & mask2]
                num_close = len(close_points)
                close_idxs = load_data.index[mask1 & mask2]

                if num_close > 0:
                    refs = [load_data.at[idx, 'reference']] + [load_data.at[i, 'reference'] for i in close_idxs] #adjoin references
                    merged_ref = "; ".join(str(r) for r in refs if pd.notna(r))

                    depth1_vals = [load_data.at[idx, 'depth1']] + [load_data.at[i, 'depth1'] for i in close_idxs]
                    merged_depth1 = np.round(np.mean(depth1_vals), 4)

                    depth2_vals = [load_data.at[idx, 'depth2']] + [load_data.at[i, 'depth2'] for i in close_idxs]
                    merged_depth2 = np.round(np.mean(depth2_vals), 4)

                    new_merged_rows.append({'reference': merged_ref, 'depth1': merged_depth1, 'depth2': merged_depth2}) #create new merged row

                    drop_rows_merge.add(idx)
                    for i in close_idxs:
                        drop_rows.append(i)
                        if drop_rows.count(i) >= num_files:
                            print(f'WARNING: Row {load_data.at[i, 'depth1']} | {load_data.at[i, 'depth2']} for {pair}. Reference {load_data.at[i, 'reference']}.')
                            print(f'Called by row {load_data.at[idx, 'depth1']} | {load_data.at[idx, 'depth2']} from reference {load_data.at[idx, 'reference']}.')

            # drop duplicate rows
            drop_rows = set(drop_rows).union(drop_rows_merge)
            load_data = load_data.drop(index=drop_rows).reset_index(drop=True)
            # add merged rows
            merged_df = pd.DataFrame(new_merged_rows)
            load_data = pd.concat([load_data, merged_df], ignore_index=True)
            load_data.drop_duplicates(subset=['depth1', 'depth2'], inplace=True)
            load_data = load_data.reset_index(drop=True)

            load_data = load_data.sort_values(by=['depth1']).reset_index(drop=True)
        
            #set up pair code stuff
            load_data[f"{pair}_code"] = [f"{pair}_{idx}" for idx in load_data.index]

            #save all the links for this pair
            all_links_total[f'{pair}'] = load_data[['depth1', 'depth2']].copy(deep=True)
            all_links_total[f'{pair}'] = all_links_total[f'{pair}'].rename(columns={
                'depth1': pair.split("-")[0],
                'depth2': pair.split("-")[1]})

            # rename to create unique columns for this pair
            load_data = load_data.rename(columns={
                'depth1': f"{pair}_{core}",
                'depth2': f"{pair}_{comparison_core}",
                'reference': f"{pair}_reference",
            })

            print(f"Processed pair {pair}, total points after merging: {len(load_data)}, ({original_rows} original total rows)")
            # append rows (block)
            big_table = pd.concat([big_table, load_data],
                                  axis=0,
                                  ignore_index=True)

#if core doesn't exist in all_links_count, add it with 0 val
for core in list_sites:
    if core not in all_links_count:
        all_links_count[core] = 0

def get_core_pair(node):
    idx, root_core = node

    root_core_row = big_table.loc[idx].dropna()
    if root_core_row.index[0].split("_")[0].split("-")[0] == root_core:
        root_core_pair = root_core_row.index[0].split("_")[0].split("-")[1]
    elif root_core_row.index[0].split("_")[0].split("-")[1] == root_core:
            root_core_pair = root_core_row.index[0].split("_")[0].split("-")[0]
    else:
        sys.exit(f"ERROR: Could not determine root core pair for root core {root_core}.")

    return root_core_pair

def readable_paths(paths):
    readable_all = []
    for root_core, path in paths:
        readable = []
        root = root_core[1]

        for i, p in enumerate(path):
            if i == 0:
                readable.append(f"{root}-{p[1]} ({p[0]})")
            else:
                readable.append(f"{path[i-1][1]}-{p[1]} ({p[0]})")
        readable_all.append(readable)

    return readable_all

def find_matches(index, core_name): # Find the matches for a given index, core_name. Outputs are two separate tuples. #TODO: comment
    match_indices = []  # empty array
    match_cores = [] # empty array

    row_vals = big_table.loc[index]
    non_zero_columns = row_vals[(row_vals.notna()) & (row_vals != 0)].index.tolist()

    #pick the active column for this core from the non-zero columns, since only 2 active columns per row
    core_columns = [col for col in non_zero_columns if col.endswith(core_name)]
    if not core_columns:
        print(f"No active columns found for core {core_name} in row {index}.")
        sys.exit()

    column_name = core_columns[0]
    core_value = big_table.at[index, column_name]  # get value of core in current row

    #below filters 
    for column in big_table.columns:  # for every column
        # same core suffix, but skip the current column
        if column.endswith(core_name): #make sure column is same core
            for idx, value in big_table[column].items(): 
                if pd.notna(value) and abs(value - core_value) <= error_margin: #if value is not NA and within error margin, add index to match_indices
                    match_indices.append(idx)

                    col_check1 = column.split("-")[0]
                    col_check2 = column.split("-")[1].split("_")[0]
                    if col_check1 != core_name:
                        matching_core = col_check1
                    elif col_check2 != core_name:
                        matching_core = col_check2
                    else:
                        print(f"ERROR: Could not determine matching core for column {column} in row {index}.")
                        sys.exit()
                    match_cores.append(matching_core)
                    
    return match_indices, match_cores

@lru_cache(maxsize=None)
def find_matches_cached(index, core_name): #cache the find_matches function to speed up calls
    return find_matches(index, core_name)

def walk_back(current_floor_number, network): #walk back to root to get path. #TODO: comment out
    if current_floor_number == 0:
        return []
    path = []
    
    current_slot = len([n for n in network if n[0] == current_floor_number])

    for floor in range(current_floor_number-1, -1, -1):
        addition_mod = 0
        floor_nodes = [n for n in network if n[0] == floor]
        for i, fnode in enumerate(floor_nodes):
            addition_mod += fnode[3]
            if addition_mod >= current_slot+1:
                path.append(fnode)
                current_slot = i
                break

    path = [(p[1], p[2]) for p in path]
    return path

def extract_numbers(existing_path):
    if existing_path == []:
        return []
    else:
        return [idx for idx, _ in existing_path]

def bfs(root_index, root_core, *, max_floor=20): #breadth-first-search #TODO: comment out
    root = (root_index, root_core)
    root_pair = (root_index, get_core_pair(root))

    queue = deque([root_pair])
    next_floor = deque()
    network = []
    current_floor_number = 0
    export_pathways = []

    with tqdm(total=max_floor, position=1, desc="BFS Floors", unit='floors', leave=False) as pbar:
        while queue:

            if current_floor_number > max_floor:
                break

            floor_width = len(queue)
            for _ in tqdm(range(floor_width), total=floor_width, position=2, desc=f'Processing Floor {current_floor_number}', unit='nodes', leave=False):
                node = queue.popleft()
                idx, core = node

                existing_path = walk_back(current_floor_number, network)

                extracted_numbers = extract_numbers(existing_path)
                
                if idx in extracted_numbers:
                    match_indices, match_cores = [], []
                    existing_path.insert(0, node) #insert at beginning to preserve order
                    existing_path.reverse() #reverse to get correct order from root to leaf
                    export_pathways.append((root, existing_path))
                elif core != root_core or current_floor_number == 0:
                    match_indices, match_cores = find_matches_cached(idx, core)
                    next_floor.extend(zip(match_indices, match_cores)) #add matches to next floor
                elif core == root_core:
                    match_indices, match_cores = [], []
                    existing_path.insert(0, node) #insert at beginning to preserve order
                    existing_path.reverse() #reverse to get correct order from root to leaf
                    export_pathways.append((root, existing_path))

                network.append((current_floor_number, idx, core, len(match_indices))) 
            
            current_floor_number += 1
            pbar.update(1)
            queue.extend(next_floor)
            next_floor.clear()
        
        pbar.n = pbar.total

    
    return export_pathways, network

@lru_cache(maxsize=None)
def bfs_cached(root_index, root_core, *, max_floor=20): #cache the find_matches function to speed up calls
    return bfs(root_index, root_core, max_floor=max_floor)

# for each row, start DFS from each core on that row
every_path = []
for index, row in tqdm(big_table.iterrows(), total=len(big_table), desc="DFS", unit="row", position=0): #for each row in big_table
    row = row.dropna()
    core1 = row.index[0].split("-")[0] #names of the cores
    core2 = row.index[0].split("-")[1].split("_")[0]

    for core in (core1, core2): #for each core

        paths, network = bfs_cached(index, core, max_floor=11) #get paths and network from BFS
        for p in paths:
            every_path.append(p)

def get_depth(core, index):
    row_vals = big_table.loc[index]
    non_zero_columns = row_vals[(row_vals.notna()) & (row_vals != 0)].index.tolist()

    core_columns = [col for col in non_zero_columns if col.endswith(core)]
    if not core_columns:
        print(f"No active columns found for core {core} in row {index}.")
        sys.exit()

    column_name = core_columns[0]
    return big_table.at[index, column_name]

def get_code(index):
    row_vals = big_table.loc[index]

    non_zero_columns = row_vals[(row_vals.notna()) & (row_vals != 0)].index.tolist()

    code_columns = [c for c in non_zero_columns if c.endswith("code")]
    if not code_columns:
        print(f"No code column found for row {index}.")
        sys.exit()

    return big_table.at[index, code_columns[0]]

def get_reference(index):
    row_vals = big_table.loc[index]

    non_zero_columns = row_vals[(row_vals.notna()) & (row_vals != 0)].index.tolist()

    reference_columns = [c for c in non_zero_columns if c.endswith("reference")]
    if not reference_columns:
        print(f"No reference column found for row {index}.")
        sys.exit()

    return big_table.at[index, reference_columns[0]]

error_networks = 0
circular_networks = 0

naughty_list = set()
nice_list = set()

for i, path in tqdm(enumerate(every_path), total=len(every_path), desc="Getting depths", unit="path"):
    
    root_core = every_path[i][0][1]
    root_index = every_path[i][0][0]

    end_node_core = every_path[i][1][-1][1]
    end_node_index = every_path[i][1][-1][0]

    all_indices = [node[0] for node in every_path[i][1]]

    if root_core != end_node_core:
        continue

    if len(all_indices) < 3:
        continue

    else:

        circular_networks += 1
        root_depth = get_depth(root_core, root_index)
        end_node_depth = get_depth(end_node_core, end_node_index)
        if abs(root_depth - end_node_depth) >= error_margin:
            error_networks += 1

            for idx in all_indices:
                if idx in nice_list:
                    continue
                else:
                    naughty_list.add(idx)

        elif abs(root_depth - end_node_depth) <= error_margin:
            
            for idx in all_indices:
                nice_list.add(idx)
                if idx in naughty_list:
                    naughty_list.remove(idx)

print(f'naughty: {len(naughty_list)}')
print(f'nice: {len(nice_list)}')

filtered = big_table.loc[big_table.index.isin(naughty_list)]
filtered.to_csv(f'/Users/quinnmackay/Desktop/filtered_output.csv', index=False)



