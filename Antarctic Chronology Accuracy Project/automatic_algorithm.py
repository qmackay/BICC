import numpy as np
import scipy
import matplotlib.pyplot as plt
import pandas as pd
import xarray as xr
import os
import yaml
import itertools
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from collections import defaultdict
import sys
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from numpy._core.numeric import indices

#set working dir
os.chdir(os.path.dirname(os.path.abspath(__file__)))

project = 'Antarctic'
project_suffix = '2plus'
input_dir = 'table_out/'

load_path = os.path.join(input_dir, project+'_'+project_suffix+'.xlsx')
loaded_excel_file = pd.read_excel(load_path, sheet_name=0)
reference_sheet = pd.read_excel(load_path, sheet_name=1, usecols="G:Q")
existing_error_sheet = pd.read_excel(load_path, sheet_name='Legend, Stats, and References')

try:
    existing_index_arr = existing_error_sheet['index'].to_numpy()
except KeyError:
    existing_index_arr = []

#parameters load
with open(f'/Users/quinnmackay/Documents/GitHub/BICC/Antarctic Chronology Accuracy Project/{project}/parameters.yml') as f:
    data = yaml.safe_load(f)
list_sites = data["list_sites"]
pairs = [f"{a}-{b}" for a, b in itertools.combinations(list_sites, 2)]

#store data
error_storage = existing_error_sheet.copy(deep=True)

#saving function
def save_excel():
    with pd.ExcelWriter(load_path, mode='a', if_sheet_exists='replace') as writer:
        error_storage.to_excel(writer, sheet_name='Legend, Stats, and References', index=False)

    print('Styling Now')
    wb = load_workbook(load_path)
    ws = wb['Legend, Stats, and References']  # or wb['SheetName']

    # column width
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15
    # fills
    grey = PatternFill(fill_type="solid", fgColor="DDDDDD")
    colors = [
        PatternFill(fill_type="solid", fgColor="E8F4FF"),
        PatternFill(fill_type="solid", fgColor="E8FFE8"),
        PatternFill(fill_type="solid", fgColor="FFF4E8"),
        PatternFill(fill_type="solid", fgColor="F3E8FF"),
    ]
    max_row = ws.max_row
    max_col = ws.max_column
    # first column grey
    for r in range(1, max_row + 1):
        ws.cell(row=r, column=1).fill = grey
    # groups of 3 columns starting from column 2
    for i, start_col in enumerate(range(2, max_col + 1, 3)):
        fill = colors[i % len(colors)]
        for c in range(start_col, min(start_col + 3, max_col + 1)):
            for r in range(1, max_row + 1):
                ws.cell(row=r, column=c).fill = fill
    wb.save(load_path)

#looping through error rows
current_err = -1
for row in range(len(loaded_excel_file)):
    
    if pd.isna(loaded_excel_file.iloc[row]['Within Row Errors']): #skip rows without errors
        continue
    if (loaded_excel_file.iloc[row]['Index'] in existing_index_arr):
        val = existing_error_sheet[f'Error Pair 1'].loc[existing_error_sheet[f'index'] == loaded_excel_file.iloc[row]['Index']].iloc[0]
        if not pd.isna(val): 
            current_err += 1
            print(f'Skipping Index {loaded_excel_file.iloc[row]["Index"]}, already completed')
            continue
    current_err += 1

    filtered_row = (loaded_excel_file.iloc[row].dropna()) #remove NaNs
    error_storage.loc[current_err, 'index'] = filtered_row['Index']

    i=1
    print('      ')
    print('##### New Error Entry #####')
    for line in filtered_row.index:
        if i%2==0: #print only every other line, separate the pairings for readability
            print('------')

        if isinstance(filtered_row[line], float):
            print(f"{line.split('.')[0]}: {np.round(filtered_row[line],4)}")
        else:
            print(f"{line}: {filtered_row[line]}")
        i+=1
    print('------')

    id_error_num = 1
    while True:
        recorded_error = input("Enter the candidate error tiepoint (or type 'Done' to end current line, or 'Exit' to quit the program): ")
        if recorded_error.lower() == 'done':
            break
        elif recorded_error.lower() == 'exit':
            print('------')

            print("Exiting and saving the program.")
            save_excel()
            sys.exit()
        elif recorded_error.upper() not in pairs:
            print('------')
            print("Invalid input. Please enter a valid site pair or type 'Done' or 'Exit'.")
            print(f"Valid site pairs are: {', '.join(pairs)}")
        elif recorded_error.upper() in pairs:
            print(f" -- Recording error tiepoint: {recorded_error.upper()}")

            col = f'Error Pair {id_error_num}' #create column if it doesn't exist
            if col not in error_storage.columns:
                error_storage[col] = pd.NA
            error_storage.at[current_err, col] = recorded_error.upper()

            col = f'Error Depths {id_error_num}' #create column if it doesn't exist
            if col not in error_storage.columns:
                error_storage[col] = pd.NA
            for u, line in enumerate(filtered_row.index): #get the depths for the recorded error
                if line not in ['Index', 'Within Row Errors']:
                    if u % 2 == 0:
                        continue
                    pair_1 = filtered_row.index[u].split('.')[0]
                    pair_2 = filtered_row.index[u+1].split('.')[0]
                    if f'{pair_1}-{pair_2}'.upper() == recorded_error.upper():
                        error_storage.at[current_err, col] = f"{np.round(filtered_row.iloc[u],3)}-{np.round(filtered_row.iloc[u+1],3)}"

            col = f'Reference {id_error_num}'
            if col not in error_storage.columns:
                error_storage[col] = pd.NA
            error_storage.at[current_err, col] = reference_sheet[f'{recorded_error.upper()} Ref'].loc[reference_sheet['Index'] == filtered_row['Index']].iloc[0]  # get the reference value from the sheet
            
            id_error_num += 1

        else:
            print(f'Something is wrong, exiting program.')
            sys.exit()

print('Completed! yay :)')
save_excel()